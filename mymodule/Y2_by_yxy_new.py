import os, linecache
from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet('My data', index=0)


f = open("C:\\T1\\#CAMLinuxCOM_2019-07-27.log",'r')
sum = 0

row = 1 #控制写入excel 的起始行号
while True:
    line = f.readline()  # 对于大文件,建议用readline
    if line !="":
        if "VI PIPE STATUS" in line:
            V = []  # 这一处的lostframe数值列表
            sum += 1
            print ("发现第 %s 处 vi pipe status" % sum) #发现 vi pipe status 标记行

            # V = []  #这一处的lostframe数值列表
            flag = True
            for i in range(0,12):

                str1 = f.readline()  #逐行读取

                if flag:

                    if i%2 != 0 and i>=3:  #大于3的单数行号进行处理

                        print("第 %s 行"%i,end=" ")

                        print(str1,end="")


                        list2 = str1.split (" ")  #字手段串分离并生成列表

                        while '' in list2:    #去除列除中的空字符
                            list2.remove ('')

                        if len(list2)>3:

                            try:
                                # print(list2[-4])  #打印出这一行的 lostframe 值
                                V.append(list2[-4]) #lostframe值加入列表
                            except:
                                pass
                        else:
                            flag = False
                            print()
                # try:
                #     sheet0.append(V)
                # except:
                #     print("数据有误")

            #选择是否添加异常行进入表格
            if flag:
                row += 1
                ws.cell(row,1).value = row - 1  # ws.cell(row,col) 操作单元格的另一种格式;
                print(V)  #打印每一处的列该值值列表
                print()
                cell1 = "G" + str(row)
                cell2 = "H" + str(row)
                cell3 = "I" + str(row)
                cell4 = "J" + str(row)
                cell5 = "K" + str(row)
                ws[cell1] = V[0]
                ws[cell2] = V[1]
                ws[cell3] = V[2]
                ws[cell4] = V[3]
                ws[cell5] = V[4]

    else:
        break

save_path = "c:\\t1\\test1.xlsx"
wb.save(save_path)
f.close()




